
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

def write_append(ws, coord, value):
    row, col = coordinate_to_tuple(coord)
    # percorre todos os merges pra achar o bloco
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            rng = merged.coord              # ex: "B11:D11"
            min_row, min_col = merged.min_row, merged.min_col
            orig = ws.cell(row=min_row, column=min_col).value or ""
            new_text = f"{orig} {value}"
            ws.unmerge_cells(rng)
            ws.cell(row=min_row, column=min_col, value=new_text)
            ws.merge_cells(rng)
            return
    # se não era merge
    orig = ws.cell(row=row, column=col).value or ""
    new_text = f"{orig} {value}"
    ws.cell(row=row, column=col, value=new_text)

def fill_batch(data_file, sheet_data, template_file, sheet_temp, mapping, prefix):
    df = pd.read_excel(data_file, sheet_name=sheet_data)
    for _, row in df.iterrows():
        wb = load_workbook(template_file)
        ws = wb[sheet_temp]
        for field, coord in mapping.items():
            val = row.get(field)
            if pd.isna(val): 
                continue
            write_append(ws, coord, val)
        nome = str(row.get('Nome','')).strip().replace(' ','_') or f'anon_{_}'
        out = f"{prefix}_{nome}.xlsx"
        wb.save(out)
        print("Gerado:", out)

if __name__ == "__main__":
    data_file     = 'clientes_organizados 1.xlsx'
    template_file = 'Proposta UP Flor De Lins.xlsx'
    sheet_temp    = 'PF'

    # mapeamento clientes → célula destino (coluna de dados → célula Excel)
    mapping_clients = {
        'Nome':              'A3',
        'CPF':               'A4',
        'Data de Nascimento': 'G4',
        'RG':                'A5',
        'Órgão Expedidor':   'G5',
        'Naturalidade':      'A6',
        'Endereço':          'A7',
        'CEP':               'I7',
        'Cidade':            'A8',
        'UF':                'F6',
        'Bairro':            'G8',
        'Telefone':          'A9',
        'Email':             'G9',
        'Estado Civil':      'G6'
    }


    mapping_socios = {
        'Nome Sócio':               'B3',
        'CPF Sócio':                'B4',
        'Data de Nascimento Sócio': 'H4',
        'RG Sócio':                 'B5',
        'Órgão Expedidor Sócio':    'H5',
        'Naturalidade Sócio':       'B6',
        'Estado Civil Sócio':       'H6',
        'Endereço Sócio':           'B7',
        'CEP Sócio':                'J7',
        'Cidade Sócio':             'B8',
        'UF Sócio':                 'G8',
        'Bairro Sócio':             'H8',
        'Telefone Sócio':           'B9',
        'Email Sócio':              'H9',
    }

    # gera para clientes
    fill_batch(
        data_file=data_file,
        sheet_data='Sheet1',
        template_file=template_file,
        sheet_temp=sheet_temp,
        mapping=mapping_clients,
        prefix='Proposta_Cliente'
    )
    # # gera para sócios
    fill_batch(
        data_file=data_file,
        sheet_name='Socios',
        template_file=template_file,
        sheet_temp=sheet_temp,
        mapping=mapping_socios,
        prefix='Proposta_Socio'
    )
