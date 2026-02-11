import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

def write_append(ws, coord, value):
    row, col = coordinate_to_tuple(coord)
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            rng = merged.coord
            orig = ws.cell(row=merged.min_row, column=merged.min_col).value or ""
            new = f"{orig} {value}"
            ws.unmerge_cells(rng)
            ws.cell(row=merged.min_row, column=merged.min_col, value=new)
            ws.merge_cells(rng)
            return
    orig = ws.cell(row=row, column=col).value or ""
    ws.cell(row=row, column=col, value=f"{orig} {value}")

def fill_combined(
    data_file, template_file,
    sheet_clients, sheet_socios, sheet_temp,
    mapping_clients, mapping_socios,
    join_key, prefix
):
    # lê dados
    df_cli = pd.read_excel(data_file, sheet_name=sheet_clients)
    df_soc = pd.read_excel(data_file, sheet_name=sheet_socios).set_index(join_key)

    for _, cli in df_cli.iterrows():
        wb = load_workbook(template_file)
        ws = wb[sheet_temp]

        # preenche cliente
        for field, coord in mapping_clients.items():
            val = cli.get(field)
            if pd.notna(val):
                write_append(ws, coord, val)

        # preenche sócio correspondente via número do apto
        apto = cli.get(join_key)
        if pd.notna(apto) and apto in df_soc.index:
            soc = df_soc.loc[apto]
            if isinstance(soc, pd.DataFrame):
                soc = soc.iloc[0]
            for field, coord in mapping_socios.items():
                val = soc.get(field)
                if pd.notna(val):
                    write_append(ws, coord, val)

        nome = str(cli.get('Nome','anon')).strip().replace(' ','_')
        wb.save(f"{prefix}_{nome}.xlsx")
        print("Gerado:", f"{prefix}_{nome}.xlsx")

if __name__=="__main__":
    data_file     = 'clientes_organizados 1.xlsx'
    template_file = 'Proposta UP Flor De Lins.xlsx'
    sheet_temp    = 'PF'

    mapping_clients = {
        'Nome':               'A3',
        'CPF':                'A4',
        'Data de Nascimento': 'G4',
        'RG':                 'A5',
        'Órgão Expedidor':    'G5',
        'Naturalidade':       'A6',
        'Endereço':           'A7',
        'CEP':                'I7',
        'Cidade':             'A8',
        'UF':                 'F6',
        'Bairro':             'G8',
        'Telefone':           'A9',
        'Email':              'G9',
        'Estado Civil':       'G6',
        'Valor total imovel': 'A21',
        'Valor comissão': 'A22',
        'Valor coodernação': 'A23',
        'Torre': 'C19'
    }

    mapping_socios = {
        'Nome':               'A11',
        'CPF':                'A12',
        'Data de Nascimento': 'G12',
        'RG':                 'A13',
        'Órgão Expedidor':    'G13',
        'Naturalidade':       'A14',
        'Endereço':           'A15',
        'CEP':                'I15',
        'Cidade':             'A16',
        'UF':                 'F14',
        'Bairro':             'G16',
        'Telefone':           'A17',
        'Email':              'G17',
        'Estado Civil':       'G14'
    }

    # agora usando a coluna exata 'Apto' como chave
    join_key = 'Apto'

    fill_combined(
        data_file=data_file,
        template_file=template_file,
        sheet_clients='Sheet1',
        sheet_socios='Socios',
        sheet_temp=sheet_temp,
        mapping_clients=mapping_clients,
        mapping_socios=mapping_socios,
        join_key=join_key,
        prefix='Proposta'
    )
